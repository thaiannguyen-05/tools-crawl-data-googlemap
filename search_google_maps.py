"""
Google Maps Business Scraper
Tự động tìm kiếm và thu thập thông tin doanh nghiệp từ Google Maps
Features:
- Multi-tab parallel processing
- Graceful shutdown with Ctrl+C (saves progress)
- Resume from last position (cursor-like)
- Query-based file naming (e.g. "bất động sản" → "batdongsan")
- Excel export
"""

import json
import asyncio
import re
import random
import signal
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass, field, asdict
from playwright.async_api import async_playwright, Page, BrowserContext, TimeoutError as PlaywrightTimeoutError

# For Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl not installed. Run: pip install openpyxl")

# For Vietnamese character removal
try:
    from unidecode import unidecode
    UNIDECODE_AVAILABLE = True
except ImportError:
    UNIDECODE_AVAILABLE = False
    # Fallback mapping for common Vietnamese characters
    VIETNAMESE_MAP = {
        'á': 'a', 'à': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
        'ă': 'a', 'ắ': 'a', 'ằ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
        'â': 'a', 'ấ': 'a', 'ầ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
        'é': 'e', 'è': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
        'ê': 'e', 'ế': 'e', 'ề': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
        'í': 'i', 'ì': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
        'ó': 'o', 'ò': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
        'ô': 'o', 'ố': 'o', 'ồ': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
        'ơ': 'o', 'ớ': 'o', 'ờ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
        'ú': 'u', 'ù': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
        'ư': 'u', 'ứ': 'u', 'ừ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
        'ý': 'y', 'ỳ': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
        'đ': 'd',
        'Á': 'A', 'À': 'A', 'Ả': 'A', 'Ã': 'A', 'Ạ': 'A',
        'Ă': 'A', 'Ắ': 'A', 'Ằ': 'A', 'Ẳ': 'A', 'Ẵ': 'A', 'Ặ': 'A',
        'Â': 'A', 'Ấ': 'A', 'Ầ': 'A', 'Ẩ': 'A', 'Ẫ': 'A', 'Ậ': 'A',
        'É': 'E', 'È': 'E', 'Ẻ': 'E', 'Ẽ': 'E', 'Ẹ': 'E',
        'Ê': 'E', 'Ế': 'E', 'Ề': 'E', 'Ể': 'E', 'Ễ': 'E', 'Ệ': 'E',
        'Í': 'I', 'Ì': 'I', 'Ỉ': 'I', 'Ĩ': 'I', 'Ị': 'I',
        'Ó': 'O', 'Ò': 'O', 'Ỏ': 'O', 'Õ': 'O', 'Ọ': 'O',
        'Ô': 'O', 'Ố': 'O', 'Ồ': 'O', 'Ổ': 'O', 'Ỗ': 'O', 'Ộ': 'O',
        'Ơ': 'O', 'Ớ': 'O', 'Ờ': 'O', 'Ở': 'O', 'Ỡ': 'O', 'Ợ': 'O',
        'Ú': 'U', 'Ù': 'U', 'Ủ': 'U', 'Ũ': 'U', 'Ụ': 'U',
        'Ư': 'U', 'Ứ': 'U', 'Ừ': 'U', 'Ử': 'U', 'Ữ': 'U', 'Ự': 'U',
        'Ý': 'Y', 'Ỳ': 'Y', 'Ỷ': 'Y', 'Ỹ': 'Y', 'Ỵ': 'Y',
        'Đ': 'D'
    }


# ===== CONFIGURATION =====
STATE_DIR = Path("crawl_state")
OUTPUT_DIR = Path("output")

# Global flags for control
shutdown_requested = False
pause_requested = False
save_requested = False


class KeyboardController:
    """
    Non-blocking keyboard listener for interactive terminal control.
    Supports: P (pause/resume), S (save), Q (quit), H (help)
    """
    
    def __init__(self):
        self.running = False
        self.thread: Optional[asyncio.Task] = None
        self._old_settings = None
        
    def _get_char_non_blocking(self) -> Optional[str]:
        """Get a character from stdin without blocking (Unix only)."""
        import sys
        import select
        
        # Check if there's input available
        if select.select([sys.stdin], [], [], 0)[0]:
            try:
                import termios
                import tty
                
                fd = sys.stdin.fileno()
                old_settings = termios.tcgetattr(fd)
                try:
                    tty.setraw(fd)
                    ch = sys.stdin.read(1)
                finally:
                    termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)
                return ch
            except (ImportError, termios.error):
                return sys.stdin.read(1)
        return None
    
    async def listen(self) -> None:
        """Listen for keyboard input in async loop."""
        global shutdown_requested, pause_requested, save_requested
        
        self.running = True
        
        while self.running:
            try:
                char = self._get_char_non_blocking()
                if char:
                    char_lower = char.lower()
                    
                    if char_lower == 'p':
                        pause_requested = not pause_requested
                        if pause_requested:
                            print("\n   ⏸️  PAUSED - Nhấn [P] để tiếp tục...")
                        else:
                            print("\n   ▶️  RESUMED - Tiếp tục crawl...")
                    
                    elif char_lower == 's':
                        save_requested = True
                        print("\n   💾 Save requested...")
                    
                    elif char_lower == 'q':
                        shutdown_requested = True
                        print("\n   🛑 Quit requested - Đang lưu và thoát...")
                        break
                    
                    elif char_lower == 'h':
                        self.print_help()
                
                await asyncio.sleep(0.1)  # Check every 100ms
                
            except Exception:
                await asyncio.sleep(0.5)
    
    def print_help(self) -> None:
        """Print help menu."""
        print("\n" + "=" * 50)
        print("   ⌨️  PHÍM TẮT ĐIỀU KHIỂN")
        print("=" * 50)
        print("   [P] - Pause/Resume crawl")
        print("   [S] - Save state ngay lập tức")
        print("   [Q] - Quit và lưu dữ liệu")
        print("   [H] - Hiện menu này")
        print("=" * 50 + "\n")
    
    def start(self, loop: asyncio.AbstractEventLoop) -> None:
        """Start the keyboard listener."""
        self.thread = loop.create_task(self.listen())
    
    def stop(self) -> None:
        """Stop the keyboard listener."""
        self.running = False
        if self.thread:
            self.thread.cancel()


def print_controls_banner() -> None:
    """Print the keyboard controls banner."""
    print("\n" + "─" * 60)
    print("   ⌨️  PHÍM TẮT: [P]ause  [S]ave  [Q]uit  [H]elp")
    print("─" * 60 + "\n")


def sanitize_query_to_filename(query: str) -> str:
    """
    Convert a query string to a valid filename.
    e.g., "bất động sản Hà Nội" -> "batdongsan_ha_noi"
    
    Args:
        query: The search query string
        
    Returns:
        A sanitized filename-safe string
    """
    # First, convert Vietnamese characters to ASCII
    if UNIDECODE_AVAILABLE:
        ascii_text = unidecode(query)
    else:
        # Fallback: use manual mapping
        ascii_text = query
        for viet_char, ascii_char in VIETNAMESE_MAP.items():
            ascii_text = ascii_text.replace(viet_char, ascii_char)
    
    # Convert to lowercase
    ascii_text = ascii_text.lower()
    
    # Replace spaces and special chars with underscore
    ascii_text = re.sub(r'[^a-z0-9]+', '_', ascii_text)
    
    # Remove leading/trailing underscores
    ascii_text = ascii_text.strip('_')
    
    # Collapse multiple underscores
    ascii_text = re.sub(r'_+', '_', ascii_text)
    
    return ascii_text or "query"


@dataclass
class CrawlState:
    """Manages the crawl state for resume functionality."""
    query: str
    filename: str
    urls: List[str] = field(default_factory=list)
    current_index: int = 0
    results: List[Dict[str, str]] = field(default_factory=list)
    last_updated: str = ""
    completed: bool = False
    
    def save(self) -> None:
        """Save current state to JSON file."""
        STATE_DIR.mkdir(exist_ok=True)
        state_file = STATE_DIR / f"{self.filename}_state.json"
        
        self.last_updated = datetime.now().isoformat()
        
        with open(state_file, 'w', encoding='utf-8') as f:
            json.dump(asdict(self), f, ensure_ascii=False, indent=2)
        
        print(f"   💾 State saved: {state_file}")
    
    @classmethod
    def load(cls, filename: str) -> Optional['CrawlState']:
        """Load state from JSON file if exists."""
        state_file = STATE_DIR / f"{filename}_state.json"
        
        if not state_file.exists():
            return None
        
        try:
            with open(state_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            return cls(
                query=data['query'],
                filename=data['filename'],
                urls=data.get('urls', []),
                current_index=data.get('current_index', 0),
                results=data.get('results', []),
                last_updated=data.get('last_updated', ''),
                completed=data.get('completed', False)
            )
        except (json.JSONDecodeError, KeyError) as e:
            print(f"   ⚠️ Error loading state: {e}")
            return None
    
    @classmethod
    def find_existing(cls, query: str) -> Optional['CrawlState']:
        """Find existing state for a query."""
        filename = sanitize_query_to_filename(query)
        return cls.load(filename)
    
    def mark_completed(self) -> None:
        """Mark this crawl as completed."""
        self.completed = True
        self.save()
    
    def delete_state_file(self) -> None:
        """Delete the state file after successful completion."""
        state_file = STATE_DIR / f"{self.filename}_state.json"
        if state_file.exists():
            state_file.unlink()
            print(f"   🗑️ State file deleted: {state_file}")


def list_saved_states() -> List[Path]:
    """List all saved state files."""
    if not STATE_DIR.exists():
        return []
    return list(STATE_DIR.glob("*_state.json"))


def export_from_state_files() -> None:
    """
    Export Excel files from all saved state files.
    Useful when crawl was interrupted and Excel wasn't exported.
    """
    state_files = list_saved_states()
    
    if not state_files:
        print("📂 Không tìm thấy state files trong crawl_state/")
        return
    
    print(f"\n📂 Tìm thấy {len(state_files)} state files:")
    for i, sf in enumerate(state_files, 1):
        print(f"   {i}. {sf.name}")
    
    print()
    
    for state_file in state_files:
        filename = state_file.stem.replace("_state", "")
        state = CrawlState.load(filename)
        
        if state and state.results:
            print(f"\n📊 Exporting {state.filename}: {len(state.results)} results")
            excel_path = save_to_excel(state.results, state.query)
            if excel_path:
                print(f"   ✅ Exported: {excel_path}")
        else:
            print(f"\n⚠️ {filename}: Không có kết quả để export")


def save_to_excel(
    results: List[Dict[str, str]],
    query: str,
    output_dir: Path = OUTPUT_DIR,
    include_query_col: bool = False,
) -> Optional[Path]:
    """
    Save crawl results to an Excel file.
    
    Args:
        results: List of business info dictionaries
        query: The search query (used for filename)
        output_dir: Output directory path
        
    Returns:
        Path to the created Excel file, or None if failed
    """
    if not OPENPYXL_AVAILABLE:
        print("   ❌ openpyxl not available. Cannot export to Excel.")
        print("   💡 Run: pip install openpyxl")
        return None
    
    if not results:
        print("   ⚠️ No results to export")
        return None
    
    # Create output directory
    output_dir.mkdir(exist_ok=True)
    
    # Generate filename from query
    filename = sanitize_query_to_filename(query)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = output_dir / f"{filename}_{timestamp}.xlsx"
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Define headers
    if include_query_col:
        headers = ["STT", "Query", "Tên", "Điện thoại", "Địa chỉ", "Website", "Giờ mở cửa"]
    else:
        headers = ["STT", "Tên", "Điện thoại", "Địa chỉ", "Website", "Giờ mở cửa"]
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row, business in enumerate(results, 2):
        ws.cell(row=row, column=1, value=row - 1).border = thin_border

        if include_query_col:
            ws.cell(row=row, column=2, value=business.get('query', '')).border = thin_border
            ws.cell(row=row, column=3, value=business.get('name', '')).border = thin_border
            ws.cell(row=row, column=4, value=business.get('phone', '')).border = thin_border
            ws.cell(row=row, column=5, value=business.get('address', '')).border = thin_border
            ws.cell(row=row, column=6, value=business.get('website', '')).border = thin_border
            ws.cell(row=row, column=7, value=business.get('opening_hours', '')).border = thin_border
        else:
            ws.cell(row=row, column=2, value=business.get('name', '')).border = thin_border
            ws.cell(row=row, column=3, value=business.get('phone', '')).border = thin_border
            ws.cell(row=row, column=4, value=business.get('address', '')).border = thin_border
            ws.cell(row=row, column=5, value=business.get('website', '')).border = thin_border
            ws.cell(row=row, column=6, value=business.get('opening_hours', '')).border = thin_border
    
    # Adjust column widths
    if include_query_col:
        column_widths = [6, 30, 40, 15, 60, 40, 30]
    else:
        column_widths = [6, 40, 15, 60, 40, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Save workbook
    wb.save(excel_path)
    print(f"   📊 Excel saved: {excel_path}")
    print(f"   📈 Total records: {len(results)}")
    
    return excel_path


def save_combined_excel(
    results_by_query: Dict[str, List[Dict[str, str]]],
    output_dir: Path = OUTPUT_DIR,
) -> Optional[Path]:
    """Save all query results into a single Excel file with a Query column."""
    combined: List[Dict[str, str]] = []

    for query, businesses in results_by_query.items():
        for business in businesses:
            row = dict(business)
            row["query"] = query
            combined.append(row)

    if not combined:
        print("   ⚠️ Không có kết quả để export (combined)")
        return None
    
    # Reuse save_to_excel with query column
    return save_to_excel(combined, query="combined", output_dir=output_dir, include_query_col=True)


class GoogleMapsScraper:
    """Scraper Google Maps sử dụng Playwright"""
    
    def __init__(self, headless: bool = False, concurrent_tabs: int = 3):
        self.headless = headless
        self.concurrent_tabs = concurrent_tabs
        self.max_scroll_attempts = 100  # Số lần scroll tối đa để load hết kết quả
        self.max_retries = 3  # Số lần retry khi timeout
    
    async def search_google_maps(self, query: str, page: Page, context: BrowserContext) -> List[Dict]:
        """
        Tìm kiếm trên Google Maps và lấy danh sách kết quả
        
        Args:
            query: Từ khóa tìm kiếm
            page: Playwright page instance
            context: Browser context for multi-tab processing
            
        Returns:
            List các kết quả business
        """
        from urllib.parse import quote_plus
        
        encoded_query = quote_plus(query)
        maps_url = f"https://www.google.com/maps/search/{encoded_query}"
        
        try:
            print(f"   🗺️  Đang truy cập Google Maps...")
            await page.goto(maps_url, wait_until="domcontentloaded", timeout=60000)
            
            # Đợi kết quả load với smart wait
            print(f"   ⏳ Đang chờ kết quả Maps load...")
            try:
                await page.wait_for_selector('div[role="feed"]', timeout=10000)
                print(f"   ✅ Đã load được danh sách kết quả")
            except:
                print(f"   ⚠️ Không tìm thấy danh sách kết quả")
                return []
            
            # Scroll để load tất cả kết quả
            results_count = await self._scroll_to_load_all(page)
            print(f"   📊 Tổng số kết quả sau khi scroll: {results_count}")
            
            # Parse tất cả kết quả với multi-tab
            businesses = await self._parse_all_results_with_tabs(page, context)
            
            return businesses
            
        except PlaywrightTimeoutError:
            print(f"   ⏱️ Timeout khi load Google Maps")
            return []
        except Exception as e:
            print(f"   ❌ Lỗi: {type(e).__name__}: {e}")
            return []
    
    async def _scroll_to_load_all(self, page: Page) -> int:
        """
        Scroll sidebar để load toàn bộ kết quả
        
        Returns:
            Số lượng kết quả hiện tại
        """
        print(f"   🔄 Đang scroll để load thêm kết quả...")
        
        # Selector cho scrollable container
        # Google Maps có thể thay đổi, thử nhiều selector
        scrollable_selectors = [
            'div[role="feed"]',
            'div.m6QErb',  # Class name có thể thay đổi
            '[aria-label*="Results"]',
        ]
        
        scrollable_elem = None
        for selector in scrollable_selectors:
            elem = await page.query_selector(selector)
            if elem:
                scrollable_elem = elem
                print(f"      ✓ Tìm thấy scrollable container: {selector}")
                break
        
        if not scrollable_elem:
            print(f"      ⚠️ Không tìm thấy scrollable container")
            return 0
        
        try:
            previous_count = 0
            no_change_attempts = 0
            
            for i in range(self.max_scroll_attempts):
                # Scroll xuống
                await page.evaluate('''
                    const scrollable = document.querySelector('div[role="feed"]');
                    if (scrollable) {
                        scrollable.scrollBy(0, scrollable.scrollHeight);
                    }
                ''')
                
                # Đợi load - giảm từ 3s xuống 1.5s
                await asyncio.sleep(1.5)
                
                # Đếm số item
                # Thử nhiều selector
                items = []
                for sel in ['a[href*="/maps/place/"]', 'div[role="article"]', 'a.hfpxzc']:
                    items = await page.query_selector_all(sel)
                    if items:
                        break
                
                current_count = len(items)
                
                if current_count > previous_count:
                    print(f"      ├─ Scroll {i+1}: {current_count} kết quả (+{current_count - previous_count})")
                    previous_count = current_count
                    no_change_attempts = 0
                else:
                    no_change_attempts += 1
                    print(f"      ├─ Scroll {i+1}: {current_count} kết quả (không tăng)")
                    
                    if no_change_attempts >= 3:
                        print(f"      └─ Đã load hết (không tăng sau 3 lần)")
                        break
            
            return previous_count
            
        except Exception as e:
            print(f"   ⚠️ Lỗi scroll: {e}")
            return 0
    
    async def _parse_all_results_with_tabs(self, page: Page, context: BrowserContext) -> List[Dict]:
        """
        Parse tất cả kết quả sử dụng multi-tab parallel processing
        
        Returns:
            List các business info
        """
        businesses = []
        
        try:
            # Thu thập tất cả URLs từ search results
            possible_selectors = [
                'a.hfpxzc',  # Link chính của mỗi business (phổ biến nhất)
                'a[href*="/maps/place/"]',  # Fallback
            ]
            
            urls = []
            used_selector = None
            
            for selector in possible_selectors:
                items = await page.query_selector_all(selector)
                if items and len(items) > 0:
                    used_selector = selector
                    print(f"   ✅ Tìm thấy {len(items)} items với selector: {selector}")
                    
                    # Extract URLs
                    for item in items:
                        href = await item.get_attribute('href')
                        if href and '/maps/place/' in href:
                            urls.append(href)
                    break
            
            if not urls:
                print(f"   ❌ Không tìm thấy business URLs!")
                # Debug: lưu HTML và screenshot
                html_content = await page.content()
                with open('debug_maps.html', 'w', encoding='utf-8') as f:
                    f.write(html_content)
                await page.screenshot(path='debug_maps.png')
                print(f"   💾 Đã lưu debug_maps.html và debug_maps.png")
                return businesses
            
            # Loại bỏ duplicates
            urls = list(dict.fromkeys(urls))
            
            # Giới hạn số lượng
            max_items = min(len(urls), 30)
            urls = urls[:max_items]
            
            print(f"   📝 Sẽ crawl {len(urls)} businesses với {self.concurrent_tabs} tabs song song")
            print(f"   💡 Multi-tab parallel processing...\n")
            
            # Process URLs in batches
            batch_size = self.concurrent_tabs
            total_processed = 0
            
            for batch_idx in range(0, len(urls), batch_size):
                batch_urls = urls[batch_idx:batch_idx + batch_size]
                batch_num = (batch_idx // batch_size) + 1
                total_batches = (len(urls) + batch_size - 1) // batch_size
                
                print(f"   🔄 Batch {batch_num}/{total_batches}: Processing {len(batch_urls)} items in parallel...")
                
                # Process batch in parallel
                tasks = [
                    self._extract_from_url(url, context, total_processed + i + 1, max_items)
                    for i, url in enumerate(batch_urls)
                ]
                
                batch_results = await asyncio.gather(*tasks, return_exceptions=True)
                
                # Collect successful results
                for result in batch_results:
                    if isinstance(result, dict) and result.get('name'):
                        businesses.append(result)
                    elif isinstance(result, Exception):
                        print(f"      ⚠️ Error in batch: {result}")
                
                total_processed += len(batch_urls)
                
                # Delay between batches với random jitter (anti-detection)
                if batch_idx + batch_size < len(urls):
                    batch_delay = 1.5 + random.uniform(0, 1)
                    await asyncio.sleep(batch_delay)
                
                print()
            
            print(f"   ✅ Đã parse thành công {len(businesses)}/{max_items} kết quả")
            return businesses
            
        except Exception as e:
            print(f"   ❌ Lỗi khi parse: {e}")
            import traceback
            traceback.print_exc()
            return businesses
    
    async def _extract_from_url(self, url: str, context: BrowserContext, index: int, total: int) -> Optional[Dict]:
        """
        Mở URL trong tab mới và extract business info với retry logic
        
        Args:
            url: Business detail URL
            context: Browser context
            index: Current index for logging
            total: Total items for logging
            
        Returns:
            Business info dict hoặc None
        """
        page = None
        
        # Retry with exponential backoff
        for attempt in range(self.max_retries):
            try:
                # Mở tab mới
                page = await context.new_page()
                
                # Stagger tab opening với random jitter để tránh bị detect
                base_delay = 0.05 * (index % self.concurrent_tabs)
                jitter = random.uniform(0, 0.1)
                await asyncio.sleep(base_delay + jitter)
                
                # Navigate với timeout tăng dần theo attempt
                timeout = 30000 * (attempt + 1)
                await page.goto(url, wait_until="domcontentloaded", timeout=timeout)
                
                # Thay vì wait networkidle, wait cho selector quan trọng
                try:
                    # Wait cho tên business xuất hiện
                    await page.wait_for_selector('h1', timeout=8000)
                except:
                    # Nếu không có h1, vẫn thử extract
                    pass
                
                # Thêm delay nhỏ với random jitter để panel load đầy đủ
                await asyncio.sleep(1 + random.uniform(0, 0.3))
                
                # Extract info
                business_info = await self._extract_from_detail_panel(page)
                
                if business_info and business_info.get('name'):
                    print(f"      ✓ [{index}/{total}] {business_info['name'][:50]}")
                    if business_info.get('phone'):
                        print(f"          📞 {business_info['phone']}")
                else:
                    print(f"      ⚠️ [{index}/{total}] Không lấy được thông tin")
                
                # Success - close page and return
                await page.close()
                return business_info
                
            except PlaywrightTimeoutError as e:
                if page:
                    await page.close()
                    page = None
                
                if attempt < self.max_retries - 1:
                    # Exponential backoff before retry
                    backoff = (2 ** attempt) + random.uniform(0, 1)
                    print(f"      🔄 [{index}/{total}] Timeout, đang retry sau {backoff:.1f}s...")
                    await asyncio.sleep(backoff)
                else:
                    print(f"      ❌ [{index}/{total}] Lỗi: Timeout sau {self.max_retries} lần thử")
                    return None
                    
            except Exception as e:
                if page:
                    await page.close()
                    page = None
                    
                print(f"      ❌ [{index}/{total}] Lỗi: {type(e).__name__}: {str(e)[:50]}")
                return None
        
        return None
    
    async def _extract_from_detail_panel(self, page: Page) -> Optional[Dict]:
        """
        Extract thông tin từ detail panel bên phải
        (Sau khi đã click vào một business)
        """
        try:
            
            # Lấy tên - nhiều selector khác nhau
            name = None
            name_selectors = [
                'h1.DUwDvf',  # Selector phổ biến nhất
                'h1.fontHeadlineLarge',
                'h1',
                'div.fontHeadlineLarge span',
                '[role="main"] h1',
            ]
            
            for selector in name_selectors:
                name_elem = await page.query_selector(selector)
                if name_elem:
                    name_text = await name_elem.inner_text()
                    name_text = name_text.strip()
                    if name_text and len(name_text) > 2:
                        name = name_text
                        break
            
            if not name:
                return None
            
            # Lấy số điện thoại - nhiều cách
            phone = None
            
            # Cách 1: Tìm button có data-item-id chứa "phone"
            phone_button = await page.query_selector('button[data-item-id*="phone"]')
            if phone_button:
                aria_label = await phone_button.get_attribute('aria-label') or ''
                phone = self._extract_phone(aria_label)
            
            # Cách 2: Tìm link tel:
            if not phone:
                tel_link = await page.query_selector('a[href^="tel:"]')
                if tel_link:
                    href = await tel_link.get_attribute('href') or ''
                    phone = self._extract_phone(href)
            
            # Cách 3: Tìm trong aria-label có "Phone"
            if not phone:
                phone_buttons = await page.query_selector_all('button[aria-label*="Phone"], button[aria-label*="Điện thoại"]')
                for btn in phone_buttons:
                    aria_label = await btn.get_attribute('aria-label') or ''
                    phone = self._extract_phone(aria_label)
                    if phone:
                        break
            
            # Cách 4: Tìm trong toàn bộ panel text
            if not phone:
                # Lấy text từ phần thông tin chi tiết
                detail_sections = await page.query_selector_all('div.rogA2c')  # Sections chứa info
                for section in detail_sections:
                    text = await section.inner_text()
                    phone = self._extract_phone(text)
                    if phone:
                        break
            
            # Lấy địa chỉ
            address = "Chưa có thông tin"
            
            # Cách 1: Từ button address
            addr_button = await page.query_selector('button[data-item-id*="address"]')
            if addr_button:
                aria_label = await addr_button.get_attribute('aria-label') or ''
                if 'Address:' in aria_label or 'Địa chỉ:' in aria_label:
                    parts = aria_label.replace('Address:', '|').replace('Địa chỉ:', '|').split('|')
                    if len(parts) > 1:
                        address = parts[1].strip()
            
            # Cách 2: Tìm trong div chứa địa chỉ (thường có class fontBodyMedium)
            if address == "Chưa có thông tin":
                addr_divs = await page.query_selector_all('div.fontBodyMedium')
                for div in addr_divs:
                    text = await div.inner_text()
                    text = text.strip()
                    # Địa chỉ thường có tên thành phố và dài hơn
                    if any(city in text for city in ['Hà Nội', 'TP.HCM', 'Đà Nẵng', 'Cần Thơ', 'Hải Phòng', 'Việt Nam']):
                        if len(text) > 15 and not any(x in text for x in ['★', 'đánh giá', 'rating', 'Mở cửa', 'Đóng cửa']):
                            address = text
                            break
            
            # Cách 3: Fallback - tìm trong toàn bộ panel
            if address == "Chưa có thông tin":
                panel_elem = await page.query_selector('[role="main"]')
                if panel_elem:
                    panel_text = await panel_elem.inner_text()
                    address = self._extract_address_from_text(panel_text)
            
            # Lấy website
            website = None
            
            # Cách 1: Tìm button có data-item-id chứa "authority" hoặc "website"
            website_button = await page.query_selector('button[data-item-id*="authority"], button[data-item-id*="website"]')
            if website_button:
                aria_label = await website_button.get_attribute('aria-label') or ''
                # Extract URL từ aria-label
                website = self._extract_website(aria_label)
            
            # Cách 2: Tìm link có href bắt đầu bằng http
            if not website:
                # Tìm trong panel chính, tránh các link internal của Google Maps
                panel_elem = await page.query_selector('[role="main"]')
                if panel_elem:
                    website_links = await panel_elem.query_selector_all('a[href^="http"]')
                    for link in website_links:
                        href = await link.get_attribute('href') or ''
                        # Loại bỏ các link của Google
                        if 'google.com' not in href and 'gstatic.com' not in href:
                            website = href
                            break
            
            # Lấy thời gian hoạt động (opening hours)
            opening_hours = None
            
            # Cách 1: Tìm button có data-item-id chứa "hours"
            hours_button = await page.query_selector('button[data-item-id*="hours"]')
            if hours_button:
                aria_label = await hours_button.get_attribute('aria-label') or ''
                opening_hours = self._extract_opening_hours(aria_label)
            
            # Cách 2: Tìm trong các div chứa thông tin giờ mở cửa
            if not opening_hours:
                # Tìm text có chứa "Open", "Closes", "Mở cửa", "Đóng cửa"
                hours_indicators = ['Open', 'Closes', 'Opens', 'Mở cửa', 'Đóng cửa', '24 hours', '24 giờ']
                all_divs = await page.query_selector_all('div.fontBodyMedium, div.fontBodySmall')
                for div in all_divs:
                    text = await div.inner_text()
                    text = text.strip()
                    if any(indicator in text for indicator in hours_indicators):
                        # Tìm thêm context xung quanh để lấy đầy đủ thông tin
                        parent = await div.evaluate_handle('el => el.parentElement')
                        if parent:
                            hours_text = await parent.as_element().inner_text()
                            hours_text = hours_text.strip()
                            if len(hours_text) > 3:
                                opening_hours = hours_text
                                break
            
            return {
                "name": name,
                "phone": phone,
                "address": address,
                "website": website,
                "opening_hours": opening_hours,
            }
            
        except Exception as e:
            print(f"         Lỗi extract detail: {e}")
            return None
    
    def _extract_address_from_text(self, text: str) -> str:
        """Extract địa chỉ từ một đoạn text dài"""
        lines = text.split('\n')
        cities = ['Hà Nội', 'TP.HCM', 'TP HCM', 'Sài Gòn', 'Đà Nẵng', 'Cần Thơ', 'Hải Phòng', 'Việt Nam']
        
        for line in lines:
            line = line.strip()
            # Tìm dòng chứa tên thành phố và đủ dài
            for city in cities:
                if city in line and len(line) > 15:
                    # Loại bỏ các prefix không cần thiết
                    if ':' in line:
                        line = line.split(':', 1)[1].strip()
                    return line
        
        return "Chưa có thông tin"
    
    def _extract_website(self, text: str) -> Optional[str]:
        """Trích xuất website URL từ text"""
        # Pattern cho URL
        url_pattern = r'https?://[^\s\"\',<>]+'
        matches = re.findall(url_pattern, text)
        
        if matches:
            url = matches[0]
            # Loại bỏ các URL của Google
            if 'google.com' not in url and 'gstatic.com' not in url:
                return url
        
        # Nếu không tìm thấy http://, thử tìm domain pattern
        domain_pattern = r'(?:www\.)?[a-zA-Z0-9-]+\.[a-zA-Z]{2,}(?:\.[a-zA-Z]{2,})?'
        domain_matches = re.findall(domain_pattern, text)
        
        if domain_matches:
            domain = domain_matches[0]
            # Thêm https:// nếu chưa có
            if not domain.startswith(('http://', 'https://')):
                return f'https://{domain}'
            return domain
        
        return None
    
    def _extract_opening_hours(self, text: str) -> Optional[str]:
        """Trích xuất thông tin giờ mở cửa từ text"""
        # Làm sạch aria-label
        # Thường có format: "Hours: Open ⋅ Closes 5 PM" hoặc "Giờ: Mở cửa ⋅ Đóng cửa 17:00"
        
        # Loại bỏ các prefix như "Hours:", "Giờ:", etc.
        cleaned = text
        for prefix in ['Hours:', 'Giờ:', 'Opening hours:', 'Thời gian mở cửa:']:
            if prefix in cleaned:
                cleaned = cleaned.split(prefix, 1)[1].strip()
        
        # Nếu có nội dung hợp lệ
        if len(cleaned) > 3:
            # Làm sạch thêm các ký tự đặc biệt
            cleaned = cleaned.replace('⋅', '•').strip()
            return cleaned
        
        return None
    
    def _extract_phone(self, text: str) -> Optional[str]:
        """Trích xuất số điện thoại từ text"""
        # Pattern cho số điện thoại Việt Nam
        patterns = [
            r'(?:\+84|84|0)[\s.-]?\d{1,4}[\s.-]?\d{3}[\s.-]?\d{3,4}',
            r'(?:\+84|84|0)\d{9,10}',
            r'\b\d{10,11}\b',
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text)
            if matches:
                # Làm sạch
                phone = re.sub(r'[^\d+]', '', matches[0])
                
                # Chuẩn hóa
                if phone.startswith('+84'):
                    phone = '0' + phone[3:]
                elif phone.startswith('84'):
                    phone = '0' + phone[2:]
                
                if 10 <= len(phone) <= 11:
                    return phone
        
        return None
    
    async def run_searches(self, queries: List[str], delay: float = 3.0) -> Dict[str, List[Dict]]:
        """
        Chạy nhiều query search trên Maps
        
        Args:
            queries: Danh sách query
            delay: Delay giữa các query
            
        Returns:
            Dict với key là query, value là list kết quả
        """
        all_results = {}
        
        async with async_playwright() as p:
            print("🌐 Đang khởi động browser...")
            
            # Launch với args tương tự batdongsan_final.py
            browser = await p.chromium.launch(
                headless=self.headless,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-web-security',
                    '--no-sandbox',
                    '--disable-setuid-sandbox',
                ]
            )
            
            # Context options
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1920, "height": 1080},
                locale="vi-VN",
                timezone_id="Asia/Ho_Chi_Minh",
            )
            
            page = await context.new_page()
            
            try:
                for i, query in enumerate(queries, 1):
                    print(f"\n🔍 [{i}/{len(queries)}] Đang search: {query}")
                    
                    # Search trên Maps với context for multi-tab
                    businesses = await self.search_google_maps(query, page, context)
                    
                    all_results[query] = businesses
                    print(f"   ✅ Tổng cộng: {len(businesses)} kết quả\n")
                    
                    # 💾 Incremental save để tránh mất data khi crash
                    temp_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    temp_file = f"temp_incremental_{temp_timestamp}.json"
                    
                    try:
                        with open(temp_file, 'w', encoding='utf-8') as f:
                            json.dump(all_results, f, ensure_ascii=False, indent=2)
                        print(f"   💾 Đã lưu tạm: {temp_file}")
                    except Exception as e:
                        print(f"   ⚠️ Không thể lưu tạm: {e}")
                    
                    # Delay với random jitter
                    if i < len(queries):
                        delay_time = delay + random.uniform(0, 2)
                        print(f"   ⏳ Chờ {delay_time:.1f}s trước query tiếp theo...")
                        await asyncio.sleep(delay_time)
            
            finally:
                await browser.close()
        
        return all_results


def save_results(results: Dict[str, List[Dict]], output_file: str, timestamp: str = "", chunk_size: int = 1000):
    """Lưu kết quả vào JSON files với timestamp prefix
    Tự động chia thành nhiều files nếu > chunk_size records
    
    Args:
        results: Kết quả scraping
        output_file: Tên file gốc
        timestamp: Timestamp để thêm vào prefix (format: YYYYMMDD_HHMMSS)
        chunk_size: Số records tối đa mỗi file (default: 1000)
    """
    # Gộp và loại trùng
    all_businesses = []
    seen_names = set()
    
    for query, businesses in results.items():
        for business in businesses:
            name = business.get("name", "")
            
            # Loại trùng theo tên
            if name and name not in seen_names:
                seen_names.add(name)
                all_businesses.append(business)
    
    total_records = len(all_businesses)
    
    # Tính số files cần thiết
    num_files = (total_records + chunk_size - 1) // chunk_size
    
    print(f"\n💾 Tổng cộng {total_records} doanh nghiệp")
    
    if num_files == 1:
        # Chỉ 1 file, lưu bình thường
        if timestamp:
            if '.' in output_file:
                name_parts = output_file.rsplit('.', 1)
                final_filename = f"{timestamp}_{name_parts[0]}.{name_parts[1]}"
            else:
                final_filename = f"{timestamp}_{output_file}"
        else:
            final_filename = output_file
        
        with open(final_filename, 'w', encoding='utf-8') as f:
            json.dump(all_businesses, f, ensure_ascii=False, indent=2)
        
        print(f"✅ Đã lưu vào: {final_filename}")
    else:
        # Nhiều files, chia thành chunks
        print(f"📦 Sẽ chia thành {num_files} files ({chunk_size} records/file)")
        
        for i in range(num_files):
            start_idx = i * chunk_size
            end_idx = min((i + 1) * chunk_size, total_records)
            chunk_data = all_businesses[start_idx:end_idx]
            
            # Tạo tên file với số thứ tự
            if timestamp:
                if '.' in output_file:
                    name_parts = output_file.rsplit('.', 1)
                    chunk_filename = f"{timestamp}_{name_parts[0]}_part{i+1:03d}.{name_parts[1]}"
                else:
                    chunk_filename = f"{timestamp}_{output_file}_part{i+1:03d}"
            else:
                if '.' in output_file:
                    name_parts = output_file.rsplit('.', 1)
                    chunk_filename = f"{name_parts[0]}_part{i+1:03d}.{name_parts[1]}"
                else:
                    chunk_filename = f"{output_file}_part{i+1:03d}"
            
            with open(chunk_filename, 'w', encoding='utf-8') as f:
                json.dump(chunk_data, f, ensure_ascii=False, indent=2)
            
            print(f"   ✓ Part {i+1}/{num_files}: {chunk_filename} ({len(chunk_data)} records)")
        
        print(f"\n✅ Đã chia và lưu thành {num_files} files")



# ===== Các hàm helper để nhập query =====

def get_queries_from_args():
    """Lấy queries từ command line"""
    import sys
    if len(sys.argv) > 1:
        return sys.argv[1:]
    return None


def parse_cli_args(argv: List[str]):
    """Parse CLI args for save mode, special commands, file input, and queries."""
    save_mode = "per_query"
    special_command = None
    file_path = None
    queries: List[str] = []

    i = 0
    while i < len(argv):
        arg = argv[i]

        if arg == "--export":
            special_command = "export"
            i += 1
            continue
        if arg == "--status":
            special_command = "status"
            i += 1
            continue
        if arg == "--file" and i + 1 < len(argv):
            file_path = argv[i + 1]
            i += 2
            continue
        if arg == "--save-mode" and i + 1 < len(argv):
            save_mode = argv[i + 1].strip().lower()
            i += 2
            continue
        if arg == "--combined":
            save_mode = "combined"
            i += 1
            continue
        if arg.startswith("--"):
            i += 1
            continue

        queries.append(arg)
        i += 1

    if save_mode not in ("per_query", "combined"):
        print(f"⚠️ save-mode không hợp lệ: {save_mode} (dùng mặc định per_query)")
        save_mode = "per_query"

    return save_mode, special_command, file_path, queries


def get_queries_from_file(file_path: str) -> List[str]:
    """Đọc queries từ file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return [line.strip() for line in f if line.strip()]
    except Exception as e:
        print(f"❌ Lỗi đọc file: {e}")
        return []


def get_queries_interactive() -> List[str]:
    """Nhập queries interactive"""
    print("\n📝 NHẬP CÁC QUERY TÌM KIẾM")
    print("\n🎯 Chọn chế độ nhập:")
    print("   1. Nhập từng query (mỗi dòng 1 query, Enter 2 lần để kết thúc)")
    print("   2. Paste tất cả queries cùng lúc (Ctrl+D hoặc Ctrl+Z để kết thúc)")
    print("=" * 60)
    
    mode = input("Chọn chế độ (1/2, Enter = 1): ").strip() or "1"
    
    if mode == "2":
        # Chế độ paste nhiều queries
        print("\n📋 Paste tất cả queries vào đây (mỗi dòng 1 query)")
        print("   Nhấn Ctrl+D (Linux/Mac) hoặc Ctrl+Z + Enter (Windows) để kết thúc\n")
        
        queries = []
        try:
            while True:
                line = input()
                if line.strip():
                    queries.append(line.strip())
        except EOFError:
            # Ctrl+D hoặc Ctrl+Z
            pass
        
        if queries:
            print(f"\n✅ Đã nhận {len(queries)} queries:")
            for i, q in enumerate(queries, 1):
                print(f"   {i}. {q}")
        
        return queries
    
    else:
        # Chế độ nhập từng query
        print("\n📝 Nhập từng query, mỗi query 1 dòng")
        print("   Nhấn Enter 2 lần liên tiếp để kết thúc\n")
        
        queries = []
        empty_count = 0
        
        while True:
            query = input(f"Query {len(queries) + 1}: ").strip()
            
            if not query:
                empty_count += 1
                if empty_count >= 2:
                    break
                continue
            
            empty_count = 0
            queries.append(query)
            print(f"   ✓ Đã thêm: {query}")
        
        return queries


async def main():
    """Hàm chính với hỗ trợ resume và graceful shutdown"""
    import sys
    global shutdown_requested, pause_requested, save_requested
    
    # ============== CẤU HÌNH (Conservative - An toàn) ==============
    HEADLESS = False  # False = hiện browser để xem process
    DELAY_BETWEEN_SEARCHES = 8  # Delay giữa các query
    CONCURRENT_TABS = 3  # Số tabs song song
    BATCH_SAVE_INTERVAL = 5  # Lưu state sau mỗi 5 items
    # ===============================================================
    
    # Setup signal handlers for graceful shutdown
    def signal_handler(signum: int, frame) -> None:
        global shutdown_requested
        print("\n\n🛑 Đang dừng crawl... Lưu dữ liệu hiện tại...")
        shutdown_requested = True
    
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)
    
    print("=" * 70)
    print("🗺️  GOOGLE MAPS BUSINESS SCRAPER")
    print("   📌 Features: Resume từ vị trí dừng | Graceful shutdown | Excel export")
    print("=" * 70)
    
    save_mode, special_command, file_path, queries_from_args = parse_cli_args(sys.argv[1:])

    # Check for special commands
    if special_command == "export":
        print("📊 Exporting Excel from saved state files...")
        export_from_state_files()
        return

    if special_command == "status":
        state_files = list_saved_states()
        if state_files:
            print(f"\n📂 Saved states ({len(state_files)}):")
            for sf in state_files:
                filename = sf.stem.replace("_state", "")
                state = CrawlState.load(filename)
                if state:
                    status = "✅ completed" if state.completed else f"⏸️ {state.current_index}/{len(state.urls)}"
                    print(f"   • {state.query}: {len(state.results)} results [{status}]")
        else:
            print("\n📂 Không có state files nào được lưu")
        return
    
    # ===== NHẬP QUERIES =====
    queries = None

    # Cách 1: Command line (positional)
    if queries_from_args:
        queries = queries_from_args
        print(f"\n✅ Đã nhận {len(queries)} queries từ command line\n")

    # Cách 2: Từ file
    if not queries and file_path:
        queries = get_queries_from_file(file_path)
        if queries:
            print(f"\n✅ Đã đọc {len(queries)} queries từ file: {file_path}\n")
    
    # Cách 3: Interactive
    if not queries:
        print("\n💡 Hướng dẫn sử dụng:")
        print("   1. Command line: python script.py \"query 1\" \"query 2\"")
        print("   2. Từ file: python script.py --file queries.txt")
        print("   3. Export Excel: python script.py --export")
        print("   4. Xem status: python script.py --status")
        print("   5. Lưu chung: python script.py --save-mode combined")
        print("   6. Interactive: nhập trực tiếp\n")
        
        use_interactive = input("Bạn có muốn nhập queries ngay? (y/n): ").lower()
        if use_interactive == 'y':
            queries = get_queries_interactive()
    
    if not queries:
        print("\n❌ Không có query để search!")
        return
    
    # Hiển thị queries
    print("\n📋 DANH SÁCH QUERIES:")
    for i, q in enumerate(queries, 1):
        print(f"   {i}. {q}")
    print()
    
    print(f"⚡ Chế độ: {CONCURRENT_TABS} tabs song song")
    print(f"⏱️  Delay: {DELAY_BETWEEN_SEARCHES}s giữa các query")
    print(f"💾 Auto-save: Sau mỗi {BATCH_SAVE_INTERVAL} items")
    print(f"💾 Chế độ lưu: {save_mode}")
    
    # Print keyboard controls
    print_controls_banner()
    
    # Initialize keyboard controller
    keyboard_controller = KeyboardController()
    
    # Khởi tạo scraper
    scraper = GoogleMapsScraper(headless=HEADLESS, concurrent_tabs=CONCURRENT_TABS)
    
    # Start keyboard listener
    loop = asyncio.get_event_loop()
    keyboard_controller.start(loop)
    
    all_results_by_query: Dict[str, List[Dict[str, str]]] = {}

    # Process each query separately for better resume support
    for query_idx, query in enumerate(queries, 1):
        if shutdown_requested:
            print("\n🛑 Đã dừng theo yêu cầu người dùng")
            break
        
        filename = sanitize_query_to_filename(query)
        print(f"\n{'='*60}")
        print(f"🔍 [{query_idx}/{len(queries)}] Query: {query}")
        print(f"   📁 Filename: {filename}")
        print(f"{'='*60}")
        
        # Check for existing state
        existing_state = CrawlState.find_existing(query)
        state: CrawlState
        
        if existing_state and not existing_state.completed:
            print(f"\n📥 Tìm thấy state trước đó:")
            print(f"   • Đã crawl: {len(existing_state.results)} kết quả")
            print(f"   • Vị trí: {existing_state.current_index}/{len(existing_state.urls)}")
            print(f"   • Cập nhật: {existing_state.last_updated}")
            
            resume_choice = input("\n   Tiếp tục từ vị trí dừng? (y/n, Enter=y): ").lower().strip()
            if resume_choice in ['', 'y', 'yes']:
                state = existing_state
                print(f"   ✅ Tiếp tục từ index {state.current_index}")
            else:
                print("   🔄 Bắt đầu lại từ đầu")
                state = CrawlState(query=query, filename=filename)
        else:
            state = CrawlState(query=query, filename=filename)
        
        # Run the crawl
        async with async_playwright() as p:
            print("\n🌐 Đang khởi động browser...")
            
            browser = await p.chromium.launch(
                headless=HEADLESS,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-web-security',
                    '--no-sandbox',
                    '--disable-setuid-sandbox',
                ]
            )
            
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1920, "height": 1080},
                locale="vi-VN",
                timezone_id="Asia/Ho_Chi_Minh",
            )
            
            page = await context.new_page()
            
            try:
                # If we don't have URLs yet, search for them
                if not state.urls:
                    print("   🗺️  Đang tìm kiếm trên Google Maps...")
                    from urllib.parse import quote_plus
                    
                    encoded_query = quote_plus(query)
                    maps_url = f"https://www.google.com/maps/search/{encoded_query}"
                    
                    await page.goto(maps_url, wait_until="domcontentloaded", timeout=60000)
                    
                    try:
                        await page.wait_for_selector('div[role="feed"]', timeout=10000)
                    except:
                        print("   ⚠️ Không tìm thấy danh sách kết quả")
                        continue
                    
                    # Scroll to load all results
                    await scraper._scroll_to_load_all(page)
                    
                    # Get all URLs
                    items = await page.query_selector_all('a.hfpxzc')
                    if not items:
                        items = await page.query_selector_all('a[href*="/maps/place/"]')
                    
                    urls = []
                    for item in items:
                        href = await item.get_attribute('href')
                        if href and '/maps/place/' in href:
                            urls.append(href)
                    
                    urls = list(dict.fromkeys(urls))  # Remove duplicates
                    state.urls = urls
                    state.save()
                    
                    print(f"   📊 Tìm thấy {len(urls)} địa điểm")
                
                # Process URLs from current_index
                total_urls = len(state.urls)
                start_index = state.current_index
                
                print(f"\n   📝 Đang crawl từ index {start_index + 1}/{total_urls}...")
                
                for idx in range(start_index, total_urls):
                    # Check for pause
                    while pause_requested and not shutdown_requested:
                        await asyncio.sleep(0.5)
                    
                    if shutdown_requested:
                        print("\n   🛑 Đang lưu state và thoát...")
                        state.save()
                        break
                    
                    # Check for manual save request
                    if save_requested:
                        state.save()
                        print(f"\n   💾 Manual save: {len(state.results)} kết quả")
                        save_requested = False
                    
                    url = state.urls[idx]
                    
                    # Extract business info
                    result = await scraper._extract_from_url(url, context, idx + 1, total_urls)
                    
                    if result and result.get('name'):
                        state.results.append(result)
                    
                    state.current_index = idx + 1
                    
                    # Save state periodically
                    if (idx + 1) % BATCH_SAVE_INTERVAL == 0:
                        state.save()
                        print(f"\n   💾 Đã lưu state ({len(state.results)} kết quả)")
                    
                    # Small delay
                    await asyncio.sleep(0.5 + random.uniform(0, 0.3))
                
                # Mark completed if finished all URLs
                if state.current_index >= total_urls and not shutdown_requested:
                    state.mark_completed()
                    print(f"\n   ✅ Hoàn thành query: {len(state.results)} kết quả")
                
            except Exception as e:
                print(f"\n   ❌ Lỗi: {type(e).__name__}: {e}")
                state.save()  # Save on error
                
            finally:
                await browser.close()
        
        # Track results by query for combined export
        if state.results:
            all_results_by_query[query] = state.results

        # Save to Excel per query (if configured)
        if save_mode == "per_query":
            if state.results:
                print(f"\n   📊 Exporting {len(state.results)} results to Excel...")
                excel_path = save_to_excel(state.results, query)
                if excel_path:
                    print(f"   ✅ Excel exported: {excel_path}")
                    if state.completed:
                        state.delete_state_file()
            else:
                print("\n   ⚠️ Không có kết quả để export")
        
        # Delay before next query
        if query_idx < len(queries) and not shutdown_requested:
            delay_time = DELAY_BETWEEN_SEARCHES + random.uniform(0, 2)
            print(f"\n   ⏳ Chờ {delay_time:.1f}s trước query tiếp theo...")
            await asyncio.sleep(delay_time)
    
    # Stop keyboard listener
    keyboard_controller.stop()
    
    # Combined export (if configured)
    if save_mode == "combined" and all_results_by_query:
        print(f"\n📊 Exporting combined results from {len(all_results_by_query)} queries...")
        combined_path = save_combined_excel(all_results_by_query)
        if combined_path:
            print(f"✅ Combined Excel exported: {combined_path}")

    print("\n" + "=" * 70)
    if shutdown_requested:
        print("🛑 ĐÃ DỪNG - Dữ liệu đã được lưu")
        print("   💡 Chạy lại script để tiếp tục từ vị trí dừng")
    else:
        print("✅ HOÀN THÀNH!")
    print(f"📊 Đã xử lý {len(queries)} queries")
    print(f"📁 Kết quả: thư mục {OUTPUT_DIR}/")
    print("=" * 70)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n\n" + "=" * 70)
        print("🛑 INTERRUPTED BY USER (Ctrl+C)")
        print("=" * 70)
        
        # Export Excel from saved states
        state_files = list_saved_states()
        if state_files:
            print(f"\n📂 Đang export {len(state_files)} state files thành Excel...")
            for state_file in state_files:
                filename = state_file.stem.replace("_state", "")
                state = CrawlState.load(filename)
                if state and state.results:
                    print(f"\n   📊 {state.query}: {len(state.results)} results")
                    excel_path = save_to_excel(state.results, state.query)
                    if excel_path:
                        print(f"   ✅ Saved: {excel_path}")
        
        print("\n💡 Chạy lại script để tiếp tục từ vị trí dừng")
        print("=" * 70)
